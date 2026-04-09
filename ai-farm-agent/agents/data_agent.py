"""
DataAgent v12 — Migrado para BaseAgent.
Especialista em Excel/dados. Usa openpyxl para gerar planilhas.
"""

import json
import os
import getpass
from agents.base_agent import BaseAgent

USERNAME = getpass.getuser()
BASE = f"C:/Users/{USERNAME}"

SYSTEM_PROMPT = (
    "Voce e o DATA AGENT — especialista senior em Excel com Python.\n"
    "Seu codigo funciona na PRIMEIRA tentativa.\n\n"
    "REGRAS:\n"
    "1. SEMPRE imports: import os, subprocess, openpyxl, etc\n"
    "2. Caminho: os.path.join(os.path.expanduser('~'), 'Desktop', 'arquivo.xlsx')\n"
    "3. Formate cabecalhos: negrito, cor azul, centralizado, bordas\n"
    "4. Auto-ajuste largura colunas\n"
    "5. Abra no final: subprocess.Popen(f'start \"\" \"{filepath}\"', shell=True)\n"
    "6. SEMPRE print() resultado\n"
    "7. UMA step com codigo COMPLETO\n\n"
    "TEMPLATE:\n"
    "import os, subprocess\n"
    "from openpyxl import Workbook\n"
    "from openpyxl.styles import Font, PatternFill, Alignment, Border, Side\n"
    "from openpyxl.utils import get_column_letter\n\n"
    "desktop = os.path.join(os.path.expanduser('~'), 'Desktop')\n"
    "filepath = os.path.join(desktop, 'planilha.xlsx')\n"
    "wb = Workbook()\n"
    "ws = wb.active\n"
    "ws.title = 'Dados'\n\n"
    "hdr_font = Font(bold=True, color='FFFFFF', size=11)\n"
    "hdr_fill = PatternFill('solid', fgColor='1F4E79')\n"
    "thin = Side(style='thin')\n"
    "border = Border(left=thin, right=thin, top=thin, bottom=thin)\n\n"
    "headers = ['Col1', 'Col2']\n"
    "for col, h in enumerate(headers, 1):\n"
    "    c = ws.cell(row=1, column=col, value=h)\n"
    "    c.font = hdr_font; c.fill = hdr_fill\n"
    "    c.alignment = Alignment(horizontal='center'); c.border = border\n\n"
    "dados = [['A', 100], ['B', 200]]\n"
    "for row in dados: ws.append(row)\n"
    "for i in range(1, len(headers)+1):\n"
    "    ws.column_dimensions[get_column_letter(i)].width = 18\n"
    "for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):\n"
    "    for cell in row: cell.border = border\n\n"
    "wb.save(filepath)\n"
    "subprocess.Popen(f'start \"\" \"{filepath}\"', shell=True)\n"
    "print(f'Criado: {filepath}')\n\n"
    'FORMATO: {"steps":[{"step":1,"description":"...","code":"codigo completo"}]}'
)


class DataAgent(BaseAgent):
    """Agente especialista em dados e Excel."""

    def __init__(self):
        super().__init__(name="DATA", system_prompt=SYSTEM_PROMPT)

    def plan(self, task, context=None):
        """Gera plano e converte steps de code para run_python."""
        task_text = self._extract_task_text(task)

        ctx = ""
        if context:
            ctx = "\nCONTEXTO: " + json.dumps(context)

        try:
            raw = self._client.message(
                model=self.model,
                system=self.system_prompt,
                user_content=f"TAREFA: {task_text}{ctx}\nJSON puro.",
                max_tokens=4000,
            )
            from core.json_validator import safe_parse
            plan = safe_parse(raw, self.model)

            steps = []
            for st in plan.get("steps", []):
                code = st.get("code", "")
                code = code.replace("{BASE}", BASE).replace("{USERNAME}", USERNAME)
                steps.append({
                    "step": st.get("step", 1),
                    "description": st.get("description", ""),
                    "action": "run_python",
                    "params": {"code": code, "description": st.get("description", "")},
                    "agent": "DATA",
                })

            self.logger.info(f"Plano: {len(steps)} steps")
            self._metrics["total_plans"] += 1
            self._metrics["successful_plans"] += 1
            return {"steps": steps, "agent": "DATA"}

        except Exception as e:
            self.logger.error(f"Erro: {e}")
            self._metrics["total_plans"] += 1
            self._metrics["failed_plans"] += 1
            return {"steps": [], "error": str(e), "agent": "DATA"}